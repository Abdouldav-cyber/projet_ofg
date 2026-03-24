"""
Module de generation de rapports pour Djembe Bank
Supporte les formats CSV, Excel (XLSX) et PDF
"""
import csv
import io
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional

from sqlalchemy.orm import Session
from sqlalchemy import func, text

from app.models import User, Account, Transaction, Tontine, SupportTicket


def get_report_data(
    db: Session,
    report_type: str,
    period: str = "monthly",
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
) -> Dict[str, Any]:
    """
    Genere les donnees du rapport selon le type demande.

    Args:
        db: Session de base de donnees
        report_type: Type de rapport (users, transactions, accounts, tontines, support)
        period: Periode (daily, weekly, monthly)
        start_date: Date de debut (defaut: debut du mois)
        end_date: Date de fin (defaut: maintenant)
    """
    now = datetime.utcnow()
    if not end_date:
        end_date = now
    if not start_date:
        if period == "daily":
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif period == "weekly":
            start_date = now - timedelta(days=7)
        else:  # monthly
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    title = ""
    headers = []
    rows = []

    if report_type == "users":
        title = f"Rapport Utilisateurs - {start_date.strftime('%d/%m/%Y')} au {end_date.strftime('%d/%m/%Y')}"
        headers = ["ID", "Email", "Prenom", "Nom", "Role", "Statut", "KYC", "Date creation"]

        users = db.query(User).filter(
            User.created_at >= start_date,
            User.created_at <= end_date
        ).order_by(User.created_at.desc()).all()

        for u in users:
            is_active = u.is_active if isinstance(u.is_active, bool) else str(u.is_active).lower() == "true"
            rows.append([
                str(u.id)[:8],
                u.email,
                u.first_name or "",
                u.last_name or "",
                u.role or "customer",
                "Actif" if is_active else "Inactif",
                u.kyc_status or "pending",
                u.created_at.strftime("%d/%m/%Y %H:%M") if u.created_at else ""
            ])

        # Statistiques resumees
        total = len(rows)
        active_count = sum(1 for r in rows if r[5] == "Actif")
        summary = {
            "total_utilisateurs": total,
            "actifs": active_count,
            "inactifs": total - active_count,
        }

    elif report_type == "transactions":
        title = f"Rapport Transactions - {start_date.strftime('%d/%m/%Y')} au {end_date.strftime('%d/%m/%Y')}"
        headers = ["ID", "De", "Vers", "Montant", "Devise", "Type", "Statut", "Reference", "Date"]

        transactions = db.query(Transaction).filter(
            Transaction.created_at >= start_date,
            Transaction.created_at <= end_date
        ).order_by(Transaction.created_at.desc()).all()

        for t in transactions:
            amount = float(t.amount) if t.amount else 0
            rows.append([
                str(t.id)[:8],
                str(t.from_account_id)[:8] if t.from_account_id else "N/A",
                str(t.to_account_id)[:8] if t.to_account_id else "N/A",
                f"{amount:,.0f}",
                t.currency or "XOF",
                t.transaction_type or "transfer",
                t.status or "pending",
                t.reference or "",
                t.created_at.strftime("%d/%m/%Y %H:%M") if t.created_at else ""
            ])

        total_volume = sum(float(t.amount) if t.amount else 0 for t in transactions)
        completed = sum(1 for t in transactions if t.status == "completed")
        summary = {
            "total_transactions": len(rows),
            "completees": completed,
            "volume_total": f"{total_volume:,.0f} XOF",
        }

    elif report_type == "accounts":
        title = f"Rapport Comptes - {start_date.strftime('%d/%m/%Y')} au {end_date.strftime('%d/%m/%Y')}"
        headers = ["ID", "Utilisateur", "Type", "IBAN", "Statut", "Date creation"]

        accounts = db.query(Account).filter(
            Account.created_at >= start_date,
            Account.created_at <= end_date
        ).order_by(Account.created_at.desc()).all()

        for a in accounts:
            rows.append([
                str(a.id)[:8],
                str(a.user_id)[:8],
                a.account_type or "",
                a.iban or "N/A",
                a.status or "active",
                a.created_at.strftime("%d/%m/%Y %H:%M") if a.created_at else ""
            ])

        active = sum(1 for a in accounts if a.status == "active")
        summary = {
            "total_comptes": len(rows),
            "actifs": active,
            "geles": sum(1 for a in accounts if a.status == "frozen"),
        }

    elif report_type == "tontines":
        title = f"Rapport Tontines - {start_date.strftime('%d/%m/%Y')} au {end_date.strftime('%d/%m/%Y')}"
        headers = ["ID", "Nom", "Montant cible", "Frequence", "Methode", "Statut", "Date creation"]

        tontines = db.query(Tontine).filter(
            Tontine.created_at >= start_date,
            Tontine.created_at <= end_date
        ).order_by(Tontine.created_at.desc()).all()

        for t in tontines:
            amount = float(t.target_amount) if t.target_amount else 0
            rows.append([
                str(t.id)[:8],
                t.name or "",
                f"{amount:,.0f}",
                t.frequency or "",
                t.distribution_method or "rotating",
                t.status or "active",
                t.created_at.strftime("%d/%m/%Y %H:%M") if t.created_at else ""
            ])

        summary = {
            "total_tontines": len(rows),
            "actives": sum(1 for t in tontines if t.status == "active"),
        }

    elif report_type == "support":
        title = f"Rapport Support - {start_date.strftime('%d/%m/%Y')} au {end_date.strftime('%d/%m/%Y')}"
        headers = ["ID", "Sujet", "Categorie", "Priorite", "Statut", "Date creation"]

        tickets = db.query(SupportTicket).filter(
            SupportTicket.created_at >= start_date,
            SupportTicket.created_at <= end_date
        ).order_by(SupportTicket.created_at.desc()).all()

        for t in tickets:
            rows.append([
                str(t.id)[:8],
                t.subject or "",
                t.category or "other",
                t.priority or "medium",
                t.status or "open",
                t.created_at.strftime("%d/%m/%Y %H:%M") if t.created_at else ""
            ])

        summary = {
            "total_tickets": len(rows),
            "ouverts": sum(1 for t in tickets if t.status == "open"),
            "resolus": sum(1 for t in tickets if t.status == "resolved"),
        }

    else:
        title = "Rapport"
        headers = []
        rows = []
        summary = {}

    return {
        "title": title,
        "headers": headers,
        "rows": rows,
        "summary": summary,
        "generated_at": now.strftime("%d/%m/%Y %H:%M:%S"),
        "period": period,
    }


def generate_csv(data: Dict[str, Any]) -> bytes:
    """Genere un fichier CSV a partir des donnees du rapport."""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")

    # Titre
    writer.writerow([data["title"]])
    writer.writerow([f"Genere le: {data['generated_at']}"])
    writer.writerow([])

    # Resume
    if data.get("summary"):
        writer.writerow(["--- Resume ---"])
        for key, value in data["summary"].items():
            writer.writerow([key.replace("_", " ").title(), value])
        writer.writerow([])

    # Headers et donnees
    writer.writerow(data["headers"])
    for row in data["rows"]:
        writer.writerow(row)

    return output.getvalue().encode("utf-8-sig")  # BOM pour Excel


def generate_excel(data: Dict[str, Any]) -> bytes:
    """Genere un fichier Excel (XLSX) a partir des donnees du rapport."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("openpyxl est requis pour l'export Excel. Installez-le avec: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport"

    # Styles
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=16, color="1F2937")
    summary_font = Font(name="Calibri", bold=True, size=11, color="374151")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    # Titre
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(data["headers"]), 1))
    title_cell = ws.cell(row=1, column=1, value=data["title"])
    title_cell.font = title_font

    ws.cell(row=2, column=1, value=f"Genere le: {data['generated_at']}")

    # Resume
    row_num = 4
    if data.get("summary"):
        for key, value in data["summary"].items():
            ws.cell(row=row_num, column=1, value=key.replace("_", " ").title()).font = summary_font
            ws.cell(row=row_num, column=2, value=str(value))
            row_num += 1
        row_num += 1

    # Headers
    for col, header in enumerate(data["headers"], 1):
        cell = ws.cell(row=row_num, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row_num += 1

    # Donnees
    for row_data in data["rows"]:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left")
        row_num += 1

    # Auto-ajuster la largeur des colonnes
    for col in range(1, len(data["headers"]) + 1):
        max_length = 0
        for row in range(1, row_num):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = min(max_length + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_pdf(data: Dict[str, Any]) -> bytes:
    """Genere un fichier PDF a partir des donnees du rapport."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except ImportError:
        raise ImportError("reportlab est requis pour l'export PDF. Installez-le avec: pip install reportlab")

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=1.5 * cm, bottomMargin=1.5 * cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=18,
        spaceAfter=12,
        textColor=colors.HexColor("#1F2937"),
    )
    subtitle_style = ParagraphStyle(
        "CustomSubtitle",
        parent=styles["Normal"],
        fontSize=10,
        spaceAfter=20,
        textColor=colors.HexColor("#6B7280"),
    )

    elements = []

    # Titre
    elements.append(Paragraph(data["title"], title_style))
    elements.append(Paragraph(f"Genere le: {data['generated_at']}", subtitle_style))

    # Resume
    if data.get("summary"):
        summary_data = [["Indicateur", "Valeur"]]
        for key, value in data["summary"].items():
            summary_data.append([key.replace("_", " ").title(), str(value)])

        summary_table = Table(summary_data, colWidths=[8 * cm, 6 * cm])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, 0), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

    # Tableau principal
    if data["headers"] and data["rows"]:
        table_data = [data["headers"]] + data["rows"]

        # Calculer largeurs
        num_cols = len(data["headers"])
        available_width = landscape(A4)[0] - 3 * cm
        col_width = available_width / num_cols

        main_table = Table(table_data, colWidths=[col_width] * num_cols, repeatRows=1)
        main_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, 0), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elements.append(main_table)

    doc.build(elements)
    return output.getvalue()
